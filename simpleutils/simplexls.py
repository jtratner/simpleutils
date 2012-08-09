try:
    import openpyxl
except ImportError:
    raise ImportError("Creating Excel files requires openpyxl, please install it!")


# get Mapping ABC for:
try:
    # python3
    from collections.abc import Mapping
except ImportError:
    # or python 2.x
    from collections import Mapping

from .simpledict import dicts_to_list, list_to_dict
import simpleutils.simplecsv as sc
from .simplesets import get_all_keys

Style = openpyxl.style.Style

def make_filename(filename=None, ext='.xlsx', prefix='', **kwargs):
    """ excel-specific wrapper for make_filename"""
    append_time = kwargs.get("append_time") or False
    return sc.make_filename(filename=filename, ext=ext, prefix=prefix,
            append_time=append_time)

def get_worksheet(workbook=None, sheet_name=None):
    """
    :returns: a tuple of (workbook, worksheet) for given workbook and sheet_name

    Parameters:

    :param workbook: a workbook object (supports both
        get_sheet_by_name and add_sheet command), if None, creates one
    :param sheet_name: the sheet to be created or returned.
        if sheet_name is None, uses default name from openpyxl
        (generally "Sheet#") 

                        """
    try:
       active_ws=workbook.get_active_sheet()
       del active_ws
    except AttributeError:
        workbook = openpyxl.Workbook()
    if sheet_name is not None:
        ws = workbook.get_sheet_by_name(sheet_name)
        if ws is None:
            ws = workbook.create_sheet(title=sheet_name)
    else:
        ws = workbook.create_sheet()
    return ws


def write_dict(
        lstofdicts,
        filename = None,
        fieldnames = None,
        col_titles = None,
        row_styles = None,
        col_styles = None,
        style_dict = None,
        sheet_name = None,
        workbook = None,
        col_sort_key = None,
        ignore_keys = None,
        default_style = None,
        **kwargs):
    """Takes a given list of dicts and converts it to an excel file,
    using similar settings to write_dict in csv.

    Special Parameters:

    :param row_styles: sequence callable by row integer
    :param col_styles: sequence callable by column integer: (where 'A' --> '0')
    :param style_dict: mapping from cell names ('A1') or cell pos((0,1)) to a
          format object
    :param workbook: either an openpyxl workbook object or a string denoting the
          desired workbook title.
    :param sheet_name: string - name of sheet in which to store data (if sheet
          name already exists, it will be overwritten, otherwise one will be
          created)
    :param col_sort_key: key function for sorting column names for eventual output
    :param col_titles: list of column titles...if len(col_titles) < list of
          fieldnames found, remaining fieldnames will be appended to column titles
    :param ignore_keys: keys to ignore in dicts (these will not be written to
          file)
    :param append_time: (kwarg only!) - append_time - if True,
          :func:`write_dict` will put the header_row as the first row (or, if
          one is not given, generate one, so be aware of this)

    for each entry, :func:`write_dict` will lookup to see if there exists a
    row_style or a column_style corresponding to that entry returns
    worksheet created (note can get workbook from worksheet by:
    worksheet.parent)
    """
    ws = get_worksheet(workbook, sheet_name)
    fieldnames, data = dicts_to_list(lstofdicts,
                                            fieldnames=fieldnames, key_sorter=col_sort_key,
                                            ignore_keys=ignore_keys)
    # coerce fieldnames into strings so they can be used as Excel headers
    del lstofdicts
    fieldnames = map(str, fieldnames)
    data = [fieldnames]+data
    print("Will write first row as header names: {!r}".format(data[0]))
    use_styles = any((row_styles, col_styles, style_dict, default_style))
    settings = dict(data = data,
              row_styles = row_styles or {},
              col_styles = col_styles or {},
              default_style = default_style or Style(),
              use_styles = use_styles,
              style_dict = style_dict or {},
              filename = make_filename(filename, **kwargs),
              sheet = ws)
    del fieldnames
    _write_spreadsheet(**settings)
    return ws



def join_styles(default_style, row_style=None, col_style=None):
    if not(row_style or col_style):
        return default_style
    if row_style is None:
        return col_style
    elif col_style is None:
        return row_style
    else:
        style = Style()
        style.__dict__.update(row_style)
        style.__dict__.update(col_style)
        return style

def _write_spreadsheet(sheet, data, row_styles, col_styles,
        style_dict, filename, default_style, use_styles, **kwargs):
    """ helper function for writing spreadsheets, better to call typed
    functions directly. Additionally, best to call this function with keyword
    arguments, so no worries about ordering, etc). Will fail if any parameter
    given is None"""
    ctr = 0
    print "Storing values",
    for i, row in enumerate(data):
        for j, cell in enumerate(row):
            ctr += 1
            print ".",
            sheet[(i,j)].value = cell
    print("Successfully stored {} cells".format(ctr))
    if use_styles:
        print "\nStoring styles",
        if not(isinstance(row_styles, Mapping)):
            row_styles = list_to_dicts(row_styles)
        if not(isinstance(col_styles, Mapping)):
            col_styles = list_to_dicts(col_styles)
        join_styles = lambda row, col: join_styles(default_style, row_style=row,
                col_style=col)
        ctr = 0
        for i in range(len(data)):
            for j in range(len(row)):
                ctr += 1
                print ".",
                sheet[(i,j)].style = (style_dict.get((i,j))
                                      or join_styles(row_styles.get(i),
                                                    col_styles.get(j)))
        print("Successfully stored style in {} cells".format(ctr))
    print("Saving workbook.")
    sheet.parent.save(filename)

def basic_read_sheet(filename=None, workbook=None, sheet_name='', preserve_styles=False):
    """ reads a sheet from a workbook or file (workbook can either be a
    workbook or worksheet).

    .. caution:: Rough function!

        This is still really rough, so be careful when using.

    """
    #TODO: make this use get_fileobject for filename
    if (filename is None and workbook is None) or (filename and workbook):
        raise ValueError("Need to provide *either* a filename or a "
                "Workbook/Worksheet object")
    if filename:
        wb = openpyxl.load_workbook(filename)
        ws = (wb.get_sheet_by_name(sheet_name) or wb.get_active_sheet())
    elif workbook:
        try:
            wb = workbook.parent
        except AttributeError:
            wb = workbook
        ws = (wb.get_sheet_by_name(sheet_name) or wb.get_active_sheet())
    max_row = ws.get_highest_row()
    max_col = ws.get_highest_column()
    default_style = Style()
    style_dict = {"default":default_style}
    if preserve_styles:
        style_dict.update(dict((cell.get_coordinate(), cell.style) for row in
                ws.rows for cell in row
                if cell is not None and cell.style != default_style))
    output = [[cell.value for cell in row] for row in ws.rows]
    #TODO: somehow check to see if first row are titles or not
    return output, style_dict
